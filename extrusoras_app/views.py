import json
import barcode
import os
import sys
import psycopg2

from django.views.generic.base import TemplateView
from openpyxl import Workbook
from django.http.response import HttpResponse

from barcode.writer import ImageWriter
from barcode.codex import Code39

from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader

from reportlab.graphics.barcode import code39, code128, code93
from reportlab.graphics.barcode import eanbc, qr, usps
from reportlab.graphics.shapes import Drawing
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import mm
from reportlab.graphics import renderPDF

from datetime import datetime, time, date
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.views.generic.edit import UpdateView
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.http import HttpResponse
from django.conf import settings

from .models.maquina import Maquina
from .models.operario import Operario
from .models.remision import Remision
from .models.producto import Producto
from .models.departamento import Departamento
from .models.medicionextrusora import MedicionExtrusora
from .models.medicionimpresion import MedicionImpresion
from .models.medicionbarnizado import MedicionBarnizado
from .models.medicionrectificado import MedicionRectificado
from .models.medicionrefilado import MedicionRefilado
from .models.medicioncorrugado import MedicionCorrugado
from .models.medicionempaque import MedicionEmpaque
from .models.medicionconversion import MedicionConversion
from .models.revision import Revision
from .models.muestro import Muestro
from .models.bobina import Bobina
from .models.caja import Caja, CajaBobina

from os import scandir, rename
from os.path import abspath, splitext

from .ajax_request import *

class Reportes(ListView):
    model = MedicionExtrusora
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/reportes.html'

class ReporteEstatus(ListView):
    model = MedicionExtrusora
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/reporte_estatus.html'

    def get_context_data(self, **kwargs):
        context = super(ReporteEstatus, self).get_context_data(**kwargs)
        context['title'] = 'Reporte de Estatus de Bobinas'
        context['bobinasRemision'] = []
        bobinasRemision = BobinaRemision.objects.filter().order_by('bob_id','-id')
        id = -1
        for bR in bobinasRemision:
            if id != bR.bob_id.bob_id:
                id = bR.bob_id.bob_id
                bobina = {}
                bobina['id'] = id
                bobina['remision'] = bR.rem_id.rem_id
                bobina['departamento'] = bR.rem_id.pro_id.dep_id.dep_descripcion
                if bR.bor_estatus == 'A':
                    bobina['estatus'] = "Asignada"
                else:
                    bobina['estatus'] = "Finalizado"
                context['bobinasRemision'].append(bobina)
        return context

class ReporteBobina(ListView):
    model = MedicionExtrusora
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/reporte_bobina.html'

    def get_context_data(self, **kwargs):
        context = super(ReporteBobina, self).get_context_data(**kwargs)
        context['title'] = 'Reporte de Bobinas'
        list_departamento = Departamento.objects.exclude(dep_bach_code__exact='')
        context['departamentos'] = list_departamento
        list_maquinas = Maquina.objects.filter()
        maquinas = ""
        for m in list_maquinas:
            maquinas = maquinas+"|" if maquinas!="" else ""
            maquinas = maquinas+str(m.dep_id.dep_id)+":"+str(m.maq_id)+":"+m.maq_encabezado
        context['maquinas'] = maquinas
        return context

class ProcesoExtrusora(ListView):
    model = MedicionExtrusora
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/extrusoras.html'

    def get_context_data(self, **kwargs):
        context = super(ProcesoExtrusora, self).get_context_data(**kwargs)
        context['title'] = 'Extrusión'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=8)
        context['operarios'] = list_operario
        list_remision = Remision.objects.filter(pro_id__dep_id=8).exclude(rem_estatus="F").order_by("rem_id")
        context['remisiones'] = list_remision
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=8).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=8)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

class ProcesoAlmacen(ListView):
    model = Bobina
    context_object_name = 'almacen'
    template_name = 'extrusoras_app/almacen.html'

    def get_context_data(self, **kwargs):
        context = super(ProcesoAlmacen, self).get_context_data(**kwargs)
        context['title'] = 'Almacén'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=9)
        context['operarios'] = list_operario
        return context

class ProcesoImpresion(ListView):
    model = MedicionImpresion
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/impresion.html'

    def get_context_data(self, **kwargs):
        dep_id = 7
        context = super(ProcesoImpresion, self).get_context_data(**kwargs)
        context['title'] = 'Impresión'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=dep_id)
        context['operarios'] = list_operario
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=dep_id).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=dep_id)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

class ProcesoBarnizado(ListView):
    model = MedicionBarnizado
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/barnizado.html'

    def get_context_data(self, **kwargs):
        dep_id = 4
        context = super(ProcesoBarnizado, self).get_context_data(**kwargs)
        context['title'] = 'Barnizado'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=dep_id)
        context['operarios'] = list_operario
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=dep_id).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=dep_id)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

class ProcesoRectificado(ListView):
    model = MedicionRectificado
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/rectificado.html'

    def get_context_data(self, **kwargs):
        dep_id = 6
        context = super(ProcesoRectificado, self).get_context_data(**kwargs)
        context['title'] = 'Rectificado'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=dep_id)
        context['operarios'] = list_operario
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=dep_id).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=dep_id)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

class ProcesoRefilado(ListView):
    model = MedicionRefilado
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/refilado.html'

    def get_context_data(self, **kwargs):
        dep_id = 5
        context = super(ProcesoRefilado, self).get_context_data(**kwargs)
        context['title'] = 'Refilado'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=dep_id)
        context['operarios'] = list_operario
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=dep_id).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=dep_id)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

class ProcesoCorrugado(ListView):
    model = MedicionCorrugado
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/corrugado.html'

    def get_context_data(self, **kwargs):
        dep_id = 3
        context = super(ProcesoCorrugado, self).get_context_data(**kwargs)
        context['title'] = 'Corrugado'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=dep_id)
        context['operarios'] = list_operario
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=dep_id).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=dep_id)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

class ProcesoConversion(ListView):
    model = MedicionConversion
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/conversion.html'

    def get_context_data(self, **kwargs):
        dep_id = 1
        context = super(ProcesoConversion, self).get_context_data(**kwargs)
        context['title'] = 'Conversión'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=dep_id)
        context['operarios'] = list_operario
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=dep_id).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=dep_id)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

class ProcesoEmpaque(ListView):
    model = MedicionEmpaque
    context_object_name = 'mediciones'
    template_name = 'extrusoras_app/empaque.html'

    def get_context_data(self, **kwargs):
        dep_id = 2
        context = super(ProcesoEmpaque, self).get_context_data(**kwargs)
        context['title'] = 'Empaque'
        list_operario = Operario.objects.filter(ope_type="O",dep_id__dep_id=dep_id)
        context['operarios'] = list_operario
        list_maquinas = Maquina.objects.filter(maq_activa=True,dep_id__dep_id=dep_id).order_by("maq_id")
        context['maquinas'] = list_maquinas
        list_operario_supervisor = Operario.objects.filter(ope_type__in=('S','C'),dep_id__dep_id=dep_id)
        context['operarios_supervisores'] = list_operario_supervisor
        return context

def imprimir_extrusora(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=int(bob))
        bobina_medida = BobinaMedida.objects.get(bob_id=int(bob), dep_id=8)
        bobina_remision = BobinaRemision.objects.get(bob_id=int(bob), rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ""
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida
        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        filename = settings.MEDIA_ROOT+'/txt_bachmaster/'+fecha+'_'+hora+'_Extrusion.txt'
        f = open(abspath(filename),'w')
        f.write("OP|LOTE|FECHA|HORA|USUARIO|IDBOBINA|MUESTRA|TESTCODE|TIPO|VALOR|MEDICION")
        muestro = Muestro.objects.filter(dep_id=8).order_by("mue_id")
        mediciones = MedicionExtrusora.objects.filter(bob_id=bob).order_by("mee_id")
        mee_calibre_frontal = ""
        mee_calibre_reverso = ""
        mee_medida = ""
        mee_encog_hor = ""
        mee_encog_ver = ""
        operario = ""
        for med in mediciones:
            mee_medida = med.mee_metro
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                valido = False
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line = bobina_remision.rem_id.rem_id + "|" + bobina.bob_lote
                    line += "|" + med.mee_fecha.strftime('%Y-%m-%d')
                    line +=  "|" + med.mee_hora.strftime('%H:%M')
                    line +=  "|" + str(med.ope_id.ope_id)
                    line +=  "|" + str(bobina.bob_id)
                    line +=  "|" + str(bobina.bob_numero)
                    line +=  "|" + mue.mue_clave
                    line +=  "|" + mue.mue_tipo
                    line +=  "|" + str(med.__getattribute__(mue.mue_campo))
                    if mue.mue_campo == "mee_calibre_frontal":
                        mee_calibre_frontal = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mee_calibre_reverso":
                        mee_calibre_reverso = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mee_encog_hor":
                        mee_encog_hor = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mee_encog_ver":
                        mee_encog_ver = str(med.__getattribute__(mue.mue_campo))
                    line +=  "|" + str(med.mee_metro)
                elif med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line = bobina_remision.rem_id.rem_id + "|" + bobina.bob_lote
                    line += "|" + med.mee_fecha.strftime('%Y-%m-%d')
                    line +=  "|" + med.mee_hora.strftime('%H:%M')
                    line +=  "|" + str(med.ope_id.ope_id)
                    line +=  "|" + str(bobina.bob_id)
                    line +=  "|" + str(bobina.bob_numero)
                    line +=  "|" + mue.mue_clave
                    line +=  "|" + mue.mue_tipo
                    line +=  "|" + 'Rechazar' if med.__getattribute__(mue.mue_campo) == "S" else "|" + 'Aprobar'
                    line +=  "|" + str(med.mee_metro)
                if valido:
                    f.write('\n' + line)
        f.close()

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)

        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')

        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Calibre frontal:", 'value': mee_calibre_frontal},
                    {'label' : "Enc. ancho plano:", 'value':mee_encog_hor},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value':operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': bobina.bob_fecha.strftime('%d/%m/%Y')},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Calibre reverso:", 'value': mee_calibre_reverso},
                    {'label' : "Enc. Dir. Maq.:", 'value': mee_encog_ver},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': bobina.bob_hora.strftime('%H:%M:%S')},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA "+str(bob))

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_etiqueta_extrusora(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="EtiquetaExtrusora.pdf"'


    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=8)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ""
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida
        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        muestro = Muestro.objects.filter(dep_id=8).order_by("mue_id")
        mediciones = MedicionExtrusora.objects.filter(bob_id=bob).order_by("mee_id")
        mee_calibre_frontal = ""
        mee_calibre_reverso = ""
        mee_medida = ""
        mee_encog_hor = ""
        mee_encog_ver = ""
        operario = ""
        for med in mediciones:
            mee_medida = med.mee_metro
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    if mue.mue_campo == "mee_calibre_frontal":
                        mee_calibre_frontal = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mee_calibre_reverso":
                        mee_calibre_reverso = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mee_encog_hor":
                        mee_encog_hor = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mee_encog_ver":
                        mee_encog_ver = str(med.__getattribute__(mue.mue_campo))

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)

        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')

        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Calibre frontal:", 'value': mee_calibre_frontal},
                    {'label' : "Enc. ancho plano:", 'value':mee_encog_hor},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value':operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': bobina.bob_fecha.strftime('%d/%m/%Y')},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Calibre reverso:", 'value': mee_calibre_reverso},
                    {'label' : "Enc. Dir. Maq.:", 'value': mee_encog_ver},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': bobina.bob_hora.strftime('%H:%M:%S')},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_impresion(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=7)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida


        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        filename = settings.MEDIA_ROOT+'/txt_bachmaster/'+fecha+'_'+hora+'_Impresion.txt'
        f = open (filename,'w')
        f.write("OP|LOTE|FECHA|HORA|USUARIO|IDBOBINA|MUESTRA|TESTCODE|TIPO|VALOR|MEDICION")
        muestro = Muestro.objects.filter(dep_id=7).order_by("mue_id")
        mediciones = MedicionImpresion.objects.filter(bob_id=bob).order_by("mei_id")
        mei_temp_infrarojo = ""
        mei_temp_tablero = ""
        mei_cat_blanca = ""
        mei_cat_color = ""
        operario = ""
        for med in mediciones:
            mei_fecha = med.mei_fecha.strftime('%Y-%m-%d')
            mei_hora = med.mei_hora.strftime('%H:%M')
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                valido = False
                if med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line = bobina_remision.rem_id.rem_id + "|" + bobina.bob_lote
                    line += "|" + med.mei_fecha.strftime('%Y-%m-%d')
                    line +=  "|" + med.mei_hora.strftime('%H:%M')
                    line +=  "|" + str(med.ope_id.ope_id)
                    line +=  "|" + str(bobina.bob_id)
                    line +=  "|" + str(bobina.bob_numero)
                    line +=  "|" + mue.mue_clave
                    line +=  "|" + mue.mue_tipo
                    line +=  "|" + str(med.__getattribute__(mue.mue_campo))
                    line +=  "|" + str(med.mei_metro)
                    if mue.mue_campo == "mei_temp_infrarojo":
                        mei_temp_infrarojo = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mei_temp_tablero":
                        mei_temp_tablero = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mei_cat_blanca":
                        mei_cat_blanca = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mei_cat_color":
                        mei_cat_color = str(med.__getattribute__(mue.mue_campo))
                if valido:
                    f.write('\n' + line)
        f.close()

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')

        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value': peso},
                    {'label' : "Catalización Blanca:", 'value': mei_cat_blanca},
                    {'label' : "Temp. Infrarrojo:", 'value':mei_temp_infrarojo},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value':operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': mei_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Catalización Color:", 'value': mei_cat_color},
                    {'label' : "Temp. Tablero:", 'value': mei_temp_tablero},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': mei_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_etiqueta_impresion(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="EtiquetaImpresion.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=7)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida


        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        muestro = Muestro.objects.filter(dep_id=7).order_by("mue_id")
        mediciones = MedicionImpresion.objects.filter(bob_id=bob).order_by("mei_id")
        mei_temp_infrarojo = ""
        mei_temp_tablero = ""
        mei_cat_blanca = ""
        mei_cat_color = ""
        operario = ""
        for med in mediciones:
            mei_fecha = med.mei_fecha.strftime('%Y-%m-%d')
            mei_hora = med.mei_hora.strftime('%H:%M')
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                if med.__getattribute__(mue.mue_campo) is not None:
                    if mue.mue_campo == "mei_temp_infrarojo":
                        mei_temp_infrarojo = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mei_temp_tablero":
                        mei_temp_tablero = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mei_cat_blanca":
                        mei_cat_blanca = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "mei_cat_color":
                        mei_cat_color = str(med.__getattribute__(mue.mue_campo))

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')

        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value': peso},
                    {'label' : "Catalización Blanca:", 'value': mei_cat_blanca},
                    {'label' : "Temp. Infrarrojo:", 'value':mei_temp_infrarojo},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value':operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': mei_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Catalización Color:", 'value': mei_cat_color},
                    {'label' : "Temp. Tablero:", 'value': mei_temp_tablero},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': mei_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_barnizado(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=int(bob))
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=4)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida

        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        filename = settings.MEDIA_ROOT+'/txt_bachmaster/'+fecha+'_'+hora+'_Barnizado.txt'
        f = open (filename,'w')
        f.write("OP|LOTE|FECHA|HORA|USUARIO|IDBOBINA|MUESTRA|TESTCODE|TIPO|VALOR|MEDICION")
        muestro = Muestro.objects.filter(dep_id=4).order_by("mue_id")
        mediciones = MedicionBarnizado.objects.filter(bob_id=bob).order_by("meb_id")
        meb_cata_barniz = ""
        meb_temp_area = ""
        meb_vis_1 = ""
        meb_vis_2 = ""
        operario = ""
        for med in mediciones:
            meb_fecha = med.meb_fecha.strftime('%Y-%m-%d')
            meb_hora = med.meb_hora.strftime('%H:%M')
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                valido = False
                line = bobina_remision.rem_id.rem_id + "|" + bobina.bob_lote
                line += "|" + med.meb_fecha.strftime('%Y-%m-%d')
                line +=  "|" + med.meb_hora.strftime('%H:%M')
                line +=  "|" + str(med.ope_id.ope_id)
                line +=  "|" + str(bobina.bob_id)
                line +=  "|" + str(bobina.bob_numero)
                line +=  "|" + mue.mue_clave
                line +=  "|" + mue.mue_tipo
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line +=  "|" + str(med.__getattribute__(mue.mue_campo))
                    if mue.mue_campo == "meb_cata_barniz":
                        meb_cata_barniz = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "meb_temp_area":
                        meb_temp_area = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "meb_vis_1":
                        meb_vis_1 = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "meb_vis_2":
                        meb_vis_2 = str(med.__getattribute__(mue.mue_campo))
                elif med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line +=  "|" + 'Aprobar' if med.__getattribute__(mue.mue_campo) == "S" else  "|" + 'Rechazar'
                line +=  "|" + str(med.meb_metro)
                if valido:
                    f.write('\n' + line)
        f.close()

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')
        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Cat. tinta barniz:", 'value': meb_cata_barniz},
                    {'label' : "Viscosidad Unidad 1:", 'value':meb_vis_1},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value': operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': meb_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Temp área:", 'value': meb_temp_area},
                    {'label' : "Viscosidad Unidad 2:", 'value': meb_vis_2},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': meb_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_etiqueta_barnizado(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="EtiquetaBarnizado.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=4)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida

        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        muestro = Muestro.objects.filter(dep_id=4).order_by("mue_id")
        mediciones = MedicionBarnizado.objects.filter(bob_id=bob).order_by("meb_id")
        meb_cata_barniz = ""
        meb_temp_area = ""
        meb_vis_1 = ""
        meb_vis_2 = ""
        operario = ""
        for med in mediciones:
            meb_fecha = med.meb_fecha.strftime('%Y-%m-%d')
            meb_hora = med.meb_hora.strftime('%H:%M')
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    if mue.mue_campo == "meb_cata_barniz":
                        meb_cata_barniz = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "meb_temp_area":
                        meb_temp_area = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "meb_vis_1":
                        meb_vis_1 = str(med.__getattribute__(mue.mue_campo))
                    elif mue.mue_campo == "meb_vis_2":
                        meb_vis_2 = str(med.__getattribute__(mue.mue_campo))

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')
        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Cat. tinta barniz:", 'value': meb_cata_barniz},
                    {'label' : "Viscosidad Unidad 1:", 'value':meb_vis_1},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value': operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': meb_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Temp área:", 'value': meb_temp_area},
                    {'label' : "Viscosidad Unidad 2:", 'value': meb_vis_2},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': meb_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_rectificado(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=6)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida

        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        filename = settings.MEDIA_ROOT+'/txt_bachmaster/'+fecha+'_'+hora+'_Rectificado.txt'
        f = open (filename,'w')
        f.write("OP|LOTE|FECHA|HORA|USUARIO|IDBOBINA|MUESTRA|TESTCODE|TIPO|VALOR|MEDICION")
        muestro = Muestro.objects.filter(dep_id=6).order_by("mue_id")
        mediciones = MedicionRectificado.objects.filter(bob_id=bob).order_by("mer_id")
        mer_medida = ""
        mer_ancho_plano = ""
        mer_apariencia = ""
        operario = ""
        for med in mediciones:
            mer_fecha = med.mer_fecha.strftime('%Y-%m-%d')
            mer_hora = med.mer_hora.strftime('%H:%M')
            mer_medida = med.mer_metro
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                valido = False
                line = bobina_remision.rem_id.rem_id + "|" + bobina.bob_lote
                line += "|" + med.mer_fecha.strftime('%Y-%m-%d')
                line +=  "|" + med.mer_hora.strftime('%H:%M')
                line +=  "|" + str(med.ope_id.ope_id)
                line +=  "|" + str(bobina.bob_id)
                line +=  "|" + str(bobina.bob_numero)
                line +=  "|" + mue.mue_clave
                line +=  "|" + mue.mue_tipo
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line +=  "|" + str(med.__getattribute__(mue.mue_campo))
                    if mue.mue_campo == "mer_ancho_plano":
                        mer_ancho_plano = str(med.__getattribute__(mue.mue_campo))
                elif med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line +=  "|" + 'Aprobar' if med.__getattribute__(mue.mue_campo) == "S" else  "|" +  'Rechazar'
                    if mue.mue_campo == "mer_apariencia":
                        mer_apariencia = 'Aprobar' if med.__getattribute__(mue.mue_campo) == "S" else 'Rechazar'
                line +=  "|" + str(med.mer_metro)
                if valido:
                    f.write('\n' + line)
        f.close()

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')
        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Ancho plano:", 'value': mer_ancho_plano},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value': operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': mer_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Apariencia:", 'value': mer_apariencia},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': mer_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_etiqueta_rectificado(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="EtiquetaRectificado.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=6)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida

        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        muestro = Muestro.objects.filter(dep_id=6).order_by("mue_id")
        mediciones = MedicionRectificado.objects.filter(bob_id=bob).order_by("mer_id")
        mer_medida = ""
        mer_ancho_plano = ""
        mer_apariencia = ""
        operario = ""
        for med in mediciones:
            mer_fecha = med.mer_fecha.strftime('%Y-%m-%d')
            mer_hora = med.mer_hora.strftime('%H:%M')
            mer_medida = med.mer_metro
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    if mue.mue_campo == "mer_ancho_plano":
                        mer_ancho_plano = str(med.__getattribute__(mue.mue_campo))
                elif med.__getattribute__(mue.mue_campo) is not None:
                    if mue.mue_campo == "mer_apariencia":
                        mer_apariencia = 'Aprobar' if med.__getattribute__(mue.mue_campo) == "S" else 'Rechazar'

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')
        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Ancho plano:", 'value': mer_ancho_plano},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value': operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': mer_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "Apariencia:", 'value': mer_apariencia},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': mer_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_refilado(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=5)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida

        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        filename = settings.MEDIA_ROOT+'/txt_bachmaster/'+fecha+'_'+hora+'_Refilado.txt'
        f = open (filename,'w')
        f.write("OP|LOTE|FECHA|HORA|USUARIO|IDBOBINA|MUESTRA|TESTCODE|TIPO|VALOR|MEDICION")
        muestro = Muestro.objects.filter(dep_id=5).order_by("mue_id")
        mediciones = MedicionRefilado.objects.filter(bob_id=bob).order_by("mef_id")
        mef_medida = ""
        mef_ancho_plano = ""
        operario = ""
        for med in mediciones:
            mef_fecha = med.mef_fecha.strftime('%Y-%m-%d')
            mef_hora = med.mef_hora.strftime('%H:%M')
            mef_medida = med.mef_metro
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                valido = False
                line = bobina_remision.rem_id.rem_id + "|" + bobina.bob_lote
                line += "|" + med.mef_fecha.strftime('%Y-%m-%d')
                line +=  "|" + med.mef_hora.strftime('%H:%M')
                line +=  "|" + str(med.ope_id.ope_id)
                line +=  "|" + str(bobina.bob_id)
                line +=  "|" + str(bobina.bob_numero)
                line +=  "|" + mue.mue_clave
                line +=  "|" + mue.mue_tipo
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line +=  "|" + str(med.__getattribute__(mue.mue_campo))
                    if mue.mue_campo == "mef_ancho_plano":
                        mef_ancho_plano = str(med.__getattribute__(mue.mue_campo))
                elif med.__getattribute__(mue.mue_campo) is not None:
                    valido = True
                    line +=  "|" + 'Aprobar' if med.__getattribute__(mue.mue_campo) == "S" else  "|" + 'Rechazar'
                line +=  "|" + str(med.mef_metro)
                if valido:
                    f.write('\n' + line)
        f.close()

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')
        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Ancho plano:", 'value': mef_ancho_plano},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value': operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': mef_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "", 'value': ""},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': mef_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_etiqueta_refilado(request, bob, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="EtiquetaRefilado.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        bobina =  Bobina.objects.get(bob_id=bob)
        bobina_medida = BobinaMedida.objects.get(bob_id=bob, dep_id=5)
        bobina_remision = BobinaRemision.objects.get(bob_id=bob, rem_id=rem)
        kore = 0
        if bobina_remision.rem_id.pro_id.pro_kore is not None:
            kore = bobina_remision.rem_id.pro_id.pro_kore
        peso = ''
        if bobina_medida.bob_peso is not None:
            peso = str(bobina_medida.bob_peso-kore)+" Kg"
        medida = ''
        if bobina_remision.rem_id.pro_id.pro_medida is not None:
            medida = bobina_remision.rem_id.pro_id.pro_medida

        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')
        hora = datetime.now().time().strftime('%H%M%S')
        muestro = Muestro.objects.filter(dep_id=5).order_by("mue_id")
        mediciones = MedicionRefilado.objects.filter(bob_id=bob).order_by("mef_id")
        mef_medida = ""
        mef_ancho_plano = ""
        operario = ""
        for med in mediciones:
            mef_fecha = med.mef_fecha.strftime('%Y-%m-%d')
            mef_hora = med.mef_hora.strftime('%H:%M')
            mef_medida = med.mef_metro
            operario = med.ope_id.ope_nombre
            for mue in muestro:
                if mue.mue_tipo != 'P' and med.__getattribute__(mue.mue_campo) is not None:
                    if mue.mue_campo == "mef_ancho_plano":
                        mef_ancho_plano = str(med.__getattribute__(mue.mue_campo))

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(bobina_remision.maq_id.maq_nombre)

        #Generando código de barras
        valor = str(bobina.bob_id)
        ean = barcode.codex.Code39(valor, barcode.writer.ImageWriter(),add_checksum=False)
        filename = settings.MEDIA_ROOT+'/bar_code/bar_'+bobina_remision.rem_id.rem_id+"_"+str(bobina.bob_id)
        ean.save(filename)
        bar_code = ImageReader(filename+".png")
        p.drawImage(bar_code, 120, 10, width=196, height=75, mask='auto')
        first_col = [{'label' : "ID bobina:", 'value':str(bobina.bob_id)},
                    {'label' : "Num bobina:", 'value':str(bobina.bob_numero)},
                    {'label' : "Lote:", 'value':str(bobina.bob_lote)},
                    {'label' : "Peso:", 'value':peso},
                    {'label' : "Ancho plano:", 'value': mef_ancho_plano},
                    {'label' : "Descripción:", 'value':bobina_remision.rem_id.pro_id.pro_descripcion},
                    {'label' : "Color:", 'value':bobina_remision.rem_id.pro_id.col_id.col_descripcion},
                    {'label' : "Metros:", 'value':str(bobina_medida.bob_metro) if bobina_medida.bob_metro is not None else ''},
                    {'label' : "Operario:", 'value': operario}]

        second_col = [{'label' : "Remisión:", 'value':bobina_remision.rem_id.rem_id},
                    {'label' : "Fecha:", 'value': mef_fecha},
                    {'label' : "Medida:", 'value': str(medida)},
                    {'label' : "", 'value': ""},
                    {'label' : "", 'value':""},
                    {'label' : "Cod. prod:", 'value':bobina_remision.rem_id.pro_id.pro_id},
                    {'label' : "Turno:", 'value': mef_hora},
                    {'label' : "Observaciones:", 'value':str(bobina.bob_observacion) if bobina.bob_observacion is not None else ''}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)
    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA BOBINA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_corrugado(request, caj, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        caja = Caja.objects.get(caj_id=caj)
        remision = Remision.objects.get(rem_id=rem)
        cajaBobina = CajaBobina.objects.filter(caj_id=caj)
        metros = 0
        lotes = []
        for c in cajaBobina:
            metros = metros + c.cbo_metros
            lotes.append(c.bob_id.bob_lote)
        #Obtenemos los valores de las Mediciones
        #fecha = datetime.now().date().strftime('%Y-%m-%d')
        first = lotes[0]
        fecha = first
        for x in lotes:
            if x != first:
                first = x
                fecha += x

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("Corrugado")

        first_col = [{'label' : "ID Caja:", 'value':str(caja.caj_nro)},
                    {'label' : "Producto:", 'value':str(remision.pro_id.pro_id)},
                    {'label' : "Color:", 'value': remision.pro_id.col_id.col_descripcion},
                    {'label' : "Impresión:", 'value': remision.pro_id.pro_descripcion},
                    {'label' : "Sticks:", 'value': str(metros)},
                    {'label' : "Material:", 'value': remision.pro_id.mat_id.mat_descripcion},
                    {'label' : "Lote:", 'value': fecha}]

        second_col = [{'label' : "Remisión:", 'value': remision.rem_id},
                    {'label' : "Cliente:", 'value': ''},
                    {'label' : "Medida:", 'value': str(remision.pro_id.pro_medida)},
                    {'label' : "", 'value':""},
                    {'label' : "Peso:", 'value': str(caja.caj_peso)+" Kg"}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA CAJA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_conversion(request, caj, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        caja = Caja.objects.get(caj_id=caj)
        remision = Remision.objects.get(rem_id=rem)
        cajaBobina = CajaBobina.objects.filter(caj_id=caj)
        metros = 0

        lotes = []
        for c in cajaBobina:
            metros = metros + c.cbo_metros
            lotes.append(c.bob_id.bob_lote)
        #Obtenemos los valores de las Mediciones
        #fecha = datetime.now().date().strftime('%Y-%m-%d')
        first = lotes[0]
        fecha = first
        for x in lotes:
            if x != first:
                first = x
                fecha += " " + x
                
        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut(caja.maq_id.maq_nombre)

        first_col = [{'label' : "ID Caja:", 'value':str(caja.caj_nro)},
                    {'label' : "Producto:", 'value':str(remision.pro_id.pro_id)},
                    {'label' : "Color:", 'value': remision.pro_id.col_id.col_descripcion},
                    {'label' : "Impresión:", 'value': remision.pro_id.pro_descripcion},
                    {'label' : "Piezas:", 'value': str(metros)},
                    {'label' : "Material:", 'value': remision.pro_id.mat_id.mat_descripcion},
                    {'label' : "Lote:", 'value': fecha}]

        second_col = [{'label' : "Remisión:", 'value': remision.rem_id},
                    {'label' : "Cliente:", 'value': ''},
                    {'label' : "Medida:", 'value': str(remision.pro_id.pro_medida)},
                    {'label' : "", 'value':""},
                    {'label' : "Peso:", 'value': str(caja.caj_peso)+" Kg"}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA CAJA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

def imprimir_empaque(request, caj, rem):
    label_color = {"r": 0.5, "g": 0.5, "b": 0.5}
    value_color = {"r": 0, "g": 0, "b": 0}

    # Create the HttpResponse object with the appropriate PDF headers.
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'filename=NoExiste.pdf"'
    # Create the PDF object, using the response object as its "file."
    p = canvas.Canvas(response, pagesize=(432,288))

    #Agregando el logo
    filename = os.path.join(settings.BASE_DIR, 'extrusoras_app/static/extrusoras_app/img/')+'logo.png'
    logo = ImageReader(filename)
    p.drawImage(logo, 10, 230, width=80, height=62, mask='auto')

    textobject = p.beginText()

    try:
        caja = Caja.objects.get(caj_id=caj)
        remision = Remision.objects.get(rem_id=rem)
        cajaBobina = CajaBobina.objects.filter(caj_id=caj)
        metros = 0
        lotes = []
        for c in cajaBobina:
            metros = metros + c.cbo_metros
            lotes.append(c.bob_id.bob_lote)
        #Obtenemos los valores de las Mediciones
        #fecha = datetime.now().date().strftime('%Y-%m-%d')
        first = lotes[0]
        fecha = first
        for x in lotes:
            if x != first:
                first = x
                fecha += " " + x
        
        #Obtenemos los valores de las Mediciones
        fecha = datetime.now().date().strftime('%Y-%m-%d')

        textobject.setFont("Helvetica", 20)
        textobject.setTextOrigin(3.5*inch, 3.5*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("Empaque")

        #Generando código de barras
        first_col = [{'label' : "ID Caja:", 'value':str(caja.caj_nro)},
                    {'label' : "Producto:", 'value':str(remision.pro_id.pro_id)},
                    {'label' : "Color:", 'value': remision.pro_id.col_id.col_descripcion},
                    {'label' : "Impresión:", 'value': remision.pro_id.pro_descripcion},
                    {'label' : "Metros:", 'value': str(metros)},
                    {'label' : "Material:", 'value': remision.pro_id.mat_id.mat_descripcion},
                    {'label' : "Lote:", 'value': fecha}]

        second_col = [{'label' : "Remisión:", 'value': remision.rem_id},
                    {'label' : "Cliente:", 'value': ''},
                    {'label' : "Medida:", 'value': str(remision.pro_id.pro_medida)},
                    {'label' : "", 'value':""},
                    {'label' : "Peso:", 'value': str(caja.caj_peso)+" Kg"}]

        textobject.setFont("Helvetica", 10)

        textobject.setTextOrigin(10, 3.1*inch)

        for field in first_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

        textobject.setTextOrigin(3.5*inch, 3.1*inch)

        for field in second_col:
            textobject.setFillColorRGB(**label_color)
            textobject.textOut(field.get("label")+" ")
            textobject.setFillColorRGB(**value_color)
            textobject.textOut(field.get("value"))
            textobject.moveCursor(0, 16)

    except ObjectDoesNotExist:
        value_color = {"r": 0, "g": 0, "b": 0}
        textobject.setFont("Helvetica", 10)
        textobject.setTextOrigin(10, 3.1*inch)
        textobject.setFillColorRGB(**value_color)
        textobject.textOut("NO EXISTE LA CAJA")

    p.drawText(textobject)
    p.showPage()
    p.save()
    return response

class ReporteBobinasExcel(TemplateView):
     
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        dep_id = request.GET.get("dep_id")
        maq_id = request.GET.get("maq_id")
        start_date=None
        end_date = None
        filterDate = False
        
        if request.GET.get("start_date")!='':
            filterDate = True
            fecha, start_time = request.GET.get("start_date").split(' ')
            yearS,monthS,dayS = fecha.split('-')
            start_date = date(int(yearS),int(monthS),int(dayS))
            start_time = start_time + ":00"
        if request.GET.get("end_date")!='':
            filterDate = True
            fecha, end_time = request.GET.get("end_date").split(' ')
            yearE,monthE,dayE = fecha.split('-')
            end_date = date(int(yearE),int(monthE),int(dayE))
            end_time = end_time + ":00"

        data = []

        q = Q(pro_id__dep_id=dep_id)

        if dep_id=="1": #Conversion
            if start_date is not None and end_date is not None:
                qdate = Q(meo_fecha__range=(start_date,end_date))
                qtime = Q(meo_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(meo_fecha__gte=start_date)
                qtime = Q(meo_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(meo_fecha__lte=end_date)
                qtime = Q(meo_hora__lte=end_time)
        elif dep_id=="2": #Empaque
            if start_date is not None and end_date is not None:
                qdate = Q(mem_fecha__range=(start_date,end_date))
                qtime = Q(mem_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(mem_fecha__gte=start_date)
                qtime = Q(mem_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(mem_fecha__lte=end_date)
                qtime = Q(mem_hora__lte=end_time)
        elif dep_id=="3": #Corrugado
            if start_date is not None and end_date is not None:
                qdate = Q(mec_fecha__range=(start_date,end_date))
                qtime = Q(mec_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(mec_fecha__gte=start_date)
                qtime = Q(mec_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(mec_fecha__lte=end_date)
                qtime = Q(mec_hora__lte=end_time)
        elif dep_id=="4": #Barnizado
            if start_date is not None and end_date is not None:
                qdate = Q(meb_fecha__range=(start_date,end_date))
                qtime = Q(meb_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(meb_fecha__gte=start_date)
                qtime = Q(meb_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(meb_fecha__lte=end_date)
                qtime = Q(meb_hora__lte=end_time)
        elif dep_id=="5": #Refilado
            if start_date is not None and end_date is not None:
                qdate = Q(mef_fecha__range=(start_date,end_date))
                qtime = Q(mef_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(mef_fecha__gte=start_date)
                qtime = Q(mef_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(mef_fecha__lte=end_date)
                qtime = Q(mef_hora__lte=end_time)
        elif dep_id=="6": #Rectificado
            if start_date is not None and end_date is not None:
                qdate = Q(mer_fecha__range=(start_date,end_date))
                qtime = Q(mer_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(mer_fecha__gte=start_date)
                qtime = Q(mer_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(mer_fecha__lte=end_date)
                qtime = Q(mer_hora__lte=end_time)
        elif dep_id=="7": #Impresion
            if start_date is not None and end_date is not None:
                qdate = Q(mei_fecha__range=(start_date,end_date))
                qtime = Q(mei_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(mei_fecha__gte=start_date)
                qtime = Q(mei_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(mei_fecha__lte=end_date)
                qtime = Q(mei_hora__lte=end_time)
        elif dep_id=="8": #Extrusion
            if start_date is not None and end_date is not None:
                qdate = Q(mee_fecha__range=(start_date,end_date))
                qtime = Q(mee_hora__range=(start_time,end_time))
            elif start_date is not None:
                qdate = Q(mee_fecha__gte=start_date)
                qtime = Q(mee_hora__gte=start_time)
            elif end_date is not None:
                qdate = Q(mee_fecha__lte=end_date)
                qtime = Q(mee_hora__lte=end_time)

        inner_remisiones = Remision.objects.filter(q)

        for remisiones in inner_remisiones:
            qbr = Q(rem_id=remisiones.rem_id)
            #qbr.add(Q(bor_estatus="F"),Q.AND)
            if maq_id != "0":
                qbr.add(Q(maq_id=maq_id),Q.AND)
            remision = remisiones.rem_id
            producto =  remisiones.pro_id.pro_id
            descripcion =  remisiones.pro_id.pro_descripcion
            inner_bobinas = BobinaRemision.objects.filter(qbr).order_by('bob_id')
            for bobinas in inner_bobinas:
                load = False
                codigo = bobinas.bob_id.bob_id
                numero = bobinas.bob_id.bob_numero
                medida = bobinas.rem_id.pro_id.pro_medida
                
                try:
                    bobina_medida = BobinaMedida.objects.get(bob_id=codigo, dep_id=dep_id)
                    kore = 0
                    peso = 0
                    metro = 0
                    if bobinas.rem_id.pro_id.pro_kore is not None:
                        kore = bobinas.rem_id.pro_id.pro_kore
                    if bobina_medida.bob_peso is not None:
                        peso = str(bobina_medida.bob_peso-kore)+" Kg"
                    if bobina_medida.bob_metro is not None:
                        metro = str(bobina_medida.bob_metro)
                except ObjectDoesNotExist:
                    kore = 0
                    peso = 0
                    metro = 0

                if dep_id=="1": #Conversion
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (meo_fecha,' ', meo_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (meo_fecha,' ', meo_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (meo_fecha,' ', meo_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (meo_fecha,' ', meo_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionConversion.objects.raw("SELECT * FROM medicionconversion WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].meo_fecha.strftime('%d/%m/%Y')
                elif dep_id=="2": #Empaque
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (mem_fecha,' ', mem_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (mem_fecha,' ', mem_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (mem_fecha,' ', mem_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (mem_fecha,' ', mem_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionEmpaque.objects.raw("SELECT * FROM medicionempaque WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].mem_fecha.strftime('%d/%m/%Y')
                elif dep_id=="3": #Corrugado
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (mec_fecha,' ', mec_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (mec_fecha,' ', mec_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (mec_fecha,' ', mec_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (mec_fecha,' ', mec_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionCorrugado.objects.raw("SELECT * FROM medicioncorrugado WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].mec_fecha.strftime('%d/%m/%Y')
                elif dep_id=="4": #Barnizado
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (meb_fecha,' ', meb_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (meb_fecha,' ', meb_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (meb_fecha,' ', mei_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (meb_fecha,' ', meb_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionBarnizado.objects.raw("SELECT * FROM medicionbarnizado WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].meb_fecha.strftime('%d/%m/%Y')
                elif dep_id=="5": #Refilado
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (mef_fecha,' ', mef_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (mef_fecha,' ', mef_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (mef_fecha,' ', mef_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (mef_fecha,' ', mef_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionRefilado.objects.raw("SELECT * FROM medicionrefilado WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].mef_fecha.strftime('%d/%m/%Y')
                elif dep_id=="6": #Rectificado
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (mer_fecha,' ', mer_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (mer_fecha,' ', mer_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (mer_fecha,' ', mer_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (mer_fecha,' ', mer_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionRectificado.objects.raw("SELECT * FROM medicionrectificado WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].mer_fecha.strftime('%d/%m/%Y')
                elif dep_id=="7": #Impresion
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (mei_fecha,' ', mei_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (mei_fecha,' ', mei_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (mei_fecha,' ', mei_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (mei_fecha,' ', mei_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionImpresion.objects.raw("SELECT * FROM medicionimpresion WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].mei_fecha.strftime('%d/%m/%Y')
                elif dep_id=="8": #Extrusión
                    qdate = ""
                    if start_date is not None and end_date is not None:
                        qdate = " AND CONCAT (mee_fecha,' ', mee_hora)>='"+str(start_date)+" "+str(start_time)+"' AND "
                        qdate = qdate + " CONCAT (mee_fecha,' ', mee_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    elif start_date is not None:
                        qdate = " AND CONCAT (mee_fecha,' ', mee_hora)>='"+str(start_date)+" "+str(start_time)+"'"
                    elif end_date is not None:
                        qdate = " AND CONCAT (mee_fecha,' ', mee_hora)<='"+str(end_date)+" "+str(end_time)+"'"
                    mediciones = MedicionExtrusora.objects.raw("SELECT * FROM medicionextrusora WHERE bob_id="+str(codigo)+qdate+" LIMIT 1;")
                    if len(list(mediciones))>0:
                        load = True
                        fecha = mediciones[0].mee_fecha.strftime('%d/%m/%Y')
                
                if load:
                    data.append({
                        "fecha": fecha,
                        "id": codigo,
                        "remision": remision,
                        "codigo": producto,
                        "descripcion": descripcion,
                        "numero": numero,
                        "peso": peso,
                        "medida": medida,
                        "metro": metro,
                    })
        
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        ws.title = "SIN AGRUPAR"
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['A1'] = 'REPORTE DE BOBINAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('A1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A3'] = 'FECHA'
        ws['B3'] = 'ID BOBINA'
        ws['C3'] = 'PEDIDO'
        ws['D3'] = 'CÓDIGO'
        ws['E3'] = '# BOBINA'
        ws['F3'] = 'PESO'
        ws['G3'] = 'MEDIDA'
        ws['H3'] = 'METRO'
        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for d in data:
            peso = "0 Kg"
            if d["peso"] != 0:
                peso = str(round(float(d["peso"][0:-3]),3))+" Kg"
            ws.cell(row=cont,column=1).value = d["fecha"]
            ws.cell(row=cont,column=2).value = d["id"]
            ws.cell(row=cont,column=3).value = d["remision"]
            ws.cell(row=cont,column=4).value = d["codigo"]
            ws.cell(row=cont,column=5).value = d["numero"]
            ws.cell(row=cont,column=6).value = peso
            ws.cell(row=cont,column=7).value = d["medida"]
            ws.cell(row=cont,column=8).value = d["metro"]
            cont = cont + 1
        ws1 = wb.create_sheet('AGRUPADO')
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws1['A1'] = 'REPORTE DE BOBINAS AGRUPADO'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws1.merge_cells('A1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws1['A3'] = 'FECHA'
        ws1['B3'] = 'PEDIDO'
        ws1['C3'] = 'CÓDIGO'
        ws1['D3'] = 'DESCRIPCIÓN'
        ws1['E3'] = 'CANTIDAD'
        ws1['F3'] = 'PESO'
        ws1['G3'] = 'MEDIDA'
        ws1['H3'] = 'METRO'
        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        pedido = data[0]["remision"]
        peso = 0
        metro = 0
        nro_bobina = 0
        totalTotalBobinas = 0
        totalTotalPeso = 0
        for d in data:
            fecha = d["fecha"]
            idBob = d["id"]
            codigo = d["codigo"]
            descripcion = d["descripcion"]
            medida = d["medida"]
            totalTotalBobinas = totalTotalBobinas + 1
            totalTotalPeso =  totalTotalPeso + peso
            if pedido!=d["remision"]:
                ws1.cell(row=cont,column=1).value = fecha
                ws1.cell(row=cont,column=2).value = pedido
                ws1.cell(row=cont,column=3).value = codigo
                ws1.cell(row=cont,column=4).value = descripcion
                ws1.cell(row=cont,column=5).value = nro_bobina
                ws1.cell(row=cont,column=6).value = str(round(peso,3)) + " Kg"
                ws1.cell(row=cont,column=7).value = medida
                ws1.cell(row=cont,column=8).value = metro
                pedido=d["remision"]
                nro_bobina = 0
                peso = 0
                if d["peso"] != 0:
                    peso = round(float(d["peso"][0:-3]),3)
                metro = 0
                cont = cont + 1
            else: #Es el mismo pedido
                nro_bobina = nro_bobina + 1
                if d["peso"] != 0:
                    peso = peso + round(float(d["peso"][0:-3]),3)
                metro = metro + float(d["metro"])
        ws1.cell(row=cont,column=1).value = fecha
        ws1.cell(row=cont,column=2).value = pedido
        ws1.cell(row=cont,column=3).value = codigo
        ws1.cell(row=cont,column=4).value = descripcion
        ws1.cell(row=cont,column=5).value = nro_bobina
        ws1.cell(row=cont,column=6).value = str(round(peso,3)) + " Kg"
        ws1.cell(row=cont,column=7).value = medida
        ws1.cell(row=cont,column=8).value = metro
        cont = cont +1
        ws1.cell(row=cont,column=1).value = "Total"
        ws1.cell(row=cont,column=5).value = totalTotalBobinas
        ws1.cell(row=cont,column=6).value = str(round(totalTotalPeso,3)) + " Kg"
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteBobinas.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response

class ReporteEstatusExcel(TemplateView):
    #Usamos el método get para generar el archivo excel 
    def get(self, request, *args, **kwargs):
        data = []
        bobinasRemision = BobinaRemision.objects.filter().order_by('bob_id','-id')
        id = -1
        for bR in bobinasRemision:
            if id != bR.bob_id.bob_id:
                id = bR.bob_id.bob_id
                bobina = {}
                bobina['id'] = id
                bobina['remision'] = bR.rem_id.rem_id
                bobina['departamento'] = bR.rem_id.pro_id.dep_id.dep_descripcion
                if bR.bor_estatus == 'A':
                    bobina['estatus'] = "Asignada"
                else:
                    bobina['estatus'] = "Finalizado"
                data.append(bobina)
        #Creamos el libro de trabajo
        wb = Workbook()
        #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
        ws = wb.active
        #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ws['A1'] = 'ESTATUS DE BOBINAS'
        #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ws.merge_cells('A1:E1')
        #Creamos los encabezados desde la celda B3 hasta la E3
        ws['A3'] = 'ID BOBINA'
        ws['B3'] = 'PEDIDO'
        ws['C3'] = 'DEPARTAMENTO'
        ws['D3'] = 'ESTATUS'
        cont=4
        #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
        for d in data:
            ws.cell(row=cont,column=1).value = d["id"]
            ws.cell(row=cont,column=2).value = d["remision"]
            ws.cell(row=cont,column=3).value = d["departamento"]
            ws.cell(row=cont,column=4).value = d["estatus"]
            cont = cont + 1
        #Establecemos el nombre del archivo
        nombre_archivo ="ReporteEstatus.xlsx"
        #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
        response = HttpResponse(content_type="application/ms-excel") 
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
